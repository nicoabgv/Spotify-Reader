import os
import json
import openpyxl # type: ignore
import calendar
import tkinter as tk
import matplotlib.pyplot as plt # type: ignore

from heapq import nlargest
from itertools import groupby
from datetime import datetime
from tkinter import filedialog
from collections import Counter, defaultdict

class SpotifyAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Spotify Analyzer")

        self.current_json_file = None
        self.json_data = None

        self.create_widgets()
        
    def create_widgets(self):
        buttons_frame = tk.Frame(self.root)
        buttons_frame.pack(pady=10)

        open_file_button = tk.Button(buttons_frame, text="Open JSON File", command=self.open_file, font=("Arial", 14))
        open_file_button.grid(row=0, column=0, padx=10)

        analyze_button = tk.Button(buttons_frame, text="Analyze", command=self.analyze, font=("Arial", 14))
        analyze_button.grid(row=0, column=1, padx=10)

        exit_button = tk.Button(buttons_frame, text="Exit", command=self.root.destroy, font=("Arial", 14))
        exit_button.grid(row=0, column=2, padx=10)

        options_frame = tk.Frame(self.root)
        options_frame.pack(pady=10)

        options_label = tk.Label(options_frame, text="Select an option:", font=("Arial", 16))
        options_label.grid(row=0, column=0, padx=10)

        self.options_var = tk.StringVar(self.root)
        self.options_var.set("Show total of songs listened")
        options_menu = tk.OptionMenu(options_frame, self.options_var,
                             "Show total of songs listened",
                             "Total time spent listened",
                             "Show first and last played songs",
                             "Show the top 5 most played songs",
                             "Show the top 5 most played artists",
                             "Show the most skipped songs", 
                             "Show the average song duration",
                             "Show daily listening patterns",
                             "Show yearly statistics",
                             "Show daily playtime statistics",
                             "Show analysis of most used devices",
                             "Show statistical graphs",
                             "Analyze playback reasons",
                             "Export statistics to Excel")
        options_menu.config(font=("Arial", 14))
        options_menu.grid(row=0, column=1, padx=10)

        results_frame = tk.Frame(self.root)
        results_frame.pack(pady=20)

        self.results_text = tk.Text(results_frame, height=20, width=50, wrap=tk.WORD, font=("Arial", 16))
        self.results_text.grid(row=0, column=0, padx=10)
        self.results_text.config(state=tk.DISABLED)

    def open_file(self):
        file_path = filedialog.askopenfilename(title="Select JSON File", filetypes=[("JSON files", "*.json")])
        if file_path:
            self.json_data = self.open_json_file(file_path)
            if self.json_data:
                results = "JSON file loaded successfully.\n"
                self.results_text.delete(1.0, tk.END)
                self.results_text.insert(tk.END, results)
                self.show_results(results)
            else:
                results = "Error loading JSON file.\n"
                self.results_text.delete(1.0, tk.END)
                self.results_text.insert(tk.END, results)
                self.show_results()

    def analyze(self):
        if not self.json_data:
            results = "No JSON file loaded. Please open a JSON file first.\n"
            self.results_text.delete(1.0, tk.END)
            self.results_text.insert(tk.END, results)
            self.show_results(results)
            return
        
        selected_option = self.options_var.get()
        self.results_text.delete(1.0, tk.END)

        if selected_option == "Show total of songs listened":
            self.show_total_songs_listened()
        elif selected_option == "Total time spent listened":
            self.show_total_time_listened()
        elif selected_option == "Show first and last played songs":
            self.show_first_and_last_played_songs()
        elif selected_option == "Show the top 5 most played songs":
            self.show_top_played_songs()
        elif selected_option == "Show the top 5 most played artists":
            self.show_top_artists()
        elif selected_option == "Show the most skipped songs":
            self.show_most_skipped_songs()
        elif selected_option == "Show the average song duration":
            self.show_avg_duration()
        elif selected_option == "Show daily listening patterns":
            self.show_daily_listening_patterns()
        elif selected_option == "Show yearly statistics":
            self.show_yearly_statistics()
        elif selected_option == "Show daily playtime statistics":
            self.show_daily_playtime_statistics()
        elif selected_option == "Show analysis of most used devices":
            self.show_most_used_devices()
        elif selected_option == "Show statistical graphs":
            self.show_graphs()
        elif selected_option == "Analyze playback reasons":
            self.analyze_playback_reasons()
        elif selected_option == "Export statistics to Excel":
            self.export_to_excel() 
        else:
            self.results_text.insert(tk.END, "Invalid option. Please choose a valid option.\n")
    
    def show_results(self, results):
        self.results_text.config(state=tk.NORMAL)
        self.results_text.delete("1.0", tk.END)
        self.results_text.insert(tk.END, results)
        self.results_text.config(state=tk.DISABLED)
        
    def handle_exception(self, exception, error_message):
        self.results_text.insert(tk.END, f"Error: {exception}\n{error_message}\n")

    def open_json_file(self, file_path):
        try:
            if not os.path.isfile(file_path):
                raise FileNotFoundError(f"The file '{file_path}' does not exist.")

            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
                self.current_json_file = file_path
                return data
        except (FileNotFoundError, json.JSONDecodeError) as e:
            self.handle_exception(e, "Error loading JSON file.")
            return None

    def show_total_songs_listened(self):
        total_songs = len(self.json_data)
        results = f"Total number of songs listened: {total_songs}\n\n"
        self.show_results(results)

    def show_total_time_listened(self):
        total_time_ms = sum(info_reproduccion["ms_played"] for info_reproduccion in self.json_data)
        total_time_sec = total_time_ms / 1000
        total_time_min, total_time_sec = divmod(total_time_sec, 60)
        total_time_hr, total_time_min = divmod(total_time_min, 60)
        total_time_days, total_time_hr = divmod(total_time_hr, 24)

        results = f"Total time spent listened: {int(total_time_days)} days, {int(total_time_hr)} hours, {int(total_time_min)} minutes, {int(total_time_sec)} seconds\n\n"
        self.show_results(results)
    
    def show_first_and_last_played_songs(self):
    
        first_song = self.json_data[0]            
        last_song = self.json_data[-1]
        first_song_info = self.format_song_info(first_song)
        last_song_info = self.format_song_info(last_song)
            
        results = f"First song played: {first_song_info}\n"
        results += f"Last song played: {last_song_info}\n"

        self.show_results(results)
   
    def format_song_info(self, song):
        title = song.get('master_metadata_track_name', 'Unknown Title')
        artist = song.get('master_metadata_album_artist_name', 'Unknown Artist')
        timestamp_str = song.get('ts', 'Unknown Timestamp')

        timestamp = datetime.strptime(timestamp_str, "%Y-%m-%dT%H:%M:%SZ")

        formatted_timestamp = timestamp.strftime("%Y-%m-%d at %H:%M:%S")

        return f'"{title}" by "{artist}" on {formatted_timestamp}\n'
    
    def show_top_items(self, data, item_name):
        played_items = defaultdict(int)
        for playback_info in self.json_data:
            item = playback_info[data] if playback_info[data] is not None else "Unknown"
            played_items[item] += 1

        top_played_items = nlargest(5, played_items.items(), key=lambda x: x[1])

        results = f"Top 5 most played {item_name}s:\n"
        for idx, (item, plays) in enumerate(top_played_items, start=1):
            results += f"{idx}. {item}, Plays: {plays}\n"
        results += "\n"
        self.show_results(results)

    def show_top_played_songs(self):
        self.show_top_items("master_metadata_track_name", "Song")

    def show_top_artists(self):
        self.show_top_items("master_metadata_album_artist_name", "Artist")

    def show_most_skipped_songs(self):
        skipped_songs = [info["master_metadata_track_name"] for info in self.json_data if info.get("skipped")]
        skipped_songs_counter = Counter(skipped_songs)
        top_skipped_songs = nlargest(5, skipped_songs_counter.items(), key=lambda x: x[1])
        
        results = f"Top 5 most skipped songs:\n"
        for idx, (song, skips) in enumerate(top_skipped_songs, start=1):
            results += f"{idx}. {song}, Skips: {skips}\n"
        results += "\n"
        self.show_results(results)
    
    def show_avg_duration(self):
        total_durations = sum(info["ms_played"] for info in self.json_data)
        avg_duration_ms = total_durations / len(self.json_data)
        avg_duration_minutes = avg_duration_ms / 60000

        results = f"Average song duration: {avg_duration_minutes:.2f} minutes\n\n"
        self.show_results(results)

    def show_daily_listening_patterns(self):
        timestamps = [datetime.strptime(info["ts"], "%Y-%m-%dT%H:%M:%SZ") for info in self.json_data]
        timestamps.sort()

        daily_patterns = defaultdict(int)
        for day, plays in groupby(timestamps, key=lambda x: x.date()):
            daily_patterns[day] += len(list(plays))

        results = "Daily listening patterns:\n"
        for day, plays in daily_patterns.items():
            results += f"{day}: {plays} plays\n"

        self.show_results(results)

    def show_yearly_statistics(self):
        years = [datetime.strptime(info.get("ts"), "%Y-%m-%dT%H:%M:%SZ").year for info in self.json_data]
        year_counter = Counter(years)

        results = "Listening statistics by year:\n"
        for year, plays in sorted(year_counter.items()):
            results += f"{year}: {plays} plays\n"
        results += "\n"
        self.show_results(results)

    def show_daily_playtime_statistics(self):
        playtimes = [info.get("ms_played", 0) for info in self.json_data]
        playtimes = [playtime / (1000 * 60 * 60) for playtime in playtimes]  

        daily_playtime = defaultdict(float)

        for day, playtimes_group in groupby(zip(self.json_data, playtimes), key=lambda x: datetime.strptime(x[0]["ts"], "%Y-%m-%dT%H:%M:%SZ").date()):
            total_playtime = sum(playtime for _, playtime in playtimes_group)
            daily_playtime[day.strftime("%A")] += total_playtime

        sorted_daily_playtime = dict(sorted(daily_playtime.items(), key=lambda x: list(calendar.day_name).index(x[0])))

        results = "Daily playtime statistics:\n"
        for day, total_playtime in sorted_daily_playtime.items():
            results += f"{day}: {total_playtime:.2f} hours\n"
        results += "\n"
        self.show_results(results)

    def show_most_used_devices(self):
        used_devices = Counter(info["platform"] for info in self.json_data)
        results = "Analysis of most used devices:\n"
        
        for idx, (device, count) in enumerate(used_devices.most_common(), start=1):
            results += f"{idx}. {device}: {count} plays\n"
        
        results += "\n"
        self.show_results(results)
        
        self.results_text.insert(tk.END, "Analysis of most used devices:\n")
        for idx, (device, count) in enumerate(used_devices.most_common(), start=1):
            self.results_text.insert(tk.END, f"{idx}. {device}: {count} plays\n")
       
    def show_graphs(self):
        graph_options = {
            "1": ("show_top_played_songs_graph", "Show Top Played Songs"),
            "2": ("show_top_artists_graph", "Show Top Artists"),
            "3": ("show_playback_time_distribution_graph", "Show Playback Time Distribution"),
            "4": ("show_playback_trends_graph", "Show Playback Trends"),
        }

        graph_options_window = tk.Toplevel(self.root)
        graph_options_window.title("Graph Options")

        label = tk.Label(graph_options_window, text="Choose a graph to display:")
        label.pack(pady=10)

        for key, (_, label) in graph_options.items():
            button = tk.Button(graph_options_window, text=label, command=lambda key=key: self.display_graph(graph_options[key]))
            button.pack()
     
    def display_graph(self, graph_option):
        graph_function_name, _ = graph_option
        graph_function = getattr(self, graph_function_name)
        graph_function()

    def show_top_played_songs_graph(self):
        played_songs = Counter(
            info["master_metadata_track_name"]
            if info["master_metadata_track_name"] is not None
            else "None"
            for info in self.json_data
        )

        top_played_songs = nlargest(5, played_songs.items(), key=lambda x: x[1])

        song_names = []
        plays_list = []

        for song, plays in top_played_songs:
            song_names.append(song)
            plays_list.append(plays)

        plt.figure(figsize=(10, 6))
        plt.barh(song_names, plays_list, color='lightblue')
        plt.xlabel('Plays')
        plt.ylabel('Songs')
        plt.title('Top Played Songs')
        plt.tight_layout()
        plt.show()

    def show_top_artists_graph(self):
        played_artists = Counter(info["master_metadata_album_artist_name"] for info in self.json_data)
        top_played_artists = nlargest(5, played_artists.items(), key=lambda x: x[1])

        artist_names = [artist for artist, _ in top_played_artists]
        plays = [plays for _, plays in top_played_artists]

        plt.figure(figsize=(10, 6))
        plt.barh(artist_names, plays, color='lightcoral')
        plt.xlabel('Plays')
        plt.ylabel('Artists')
        plt.title('Top Played Artists')
        plt.tight_layout()
        plt.show()
        
    def show_playback_time_distribution_graph(self):
        playback_times = [info["ms_played"] / 1000 for info in self.json_data]

        plt.figure(figsize=(10, 6))
        plt.hist(playback_times, bins=20, color='green', alpha=0.7)
        plt.xlabel('Playback Time (seconds)')
        plt.ylabel('Number of Songs')
        plt.title('Distribution of Playback Times for Songs')
        plt.tight_layout()
        plt.show()

    def show_playback_trends_graph(self):
        plays_per_date = defaultdict(int)
        for info in self.json_data:
            ts = datetime.strptime(info["ts"], '%Y-%m-%dT%H:%M:%SZ')
            date = ts.date()
            plays_per_date[date] += 1

        dates = list(plays_per_date.keys())
        plays = [plays_per_date[date] for date in dates]

        plt.figure(figsize=(10, 6))
        plt.plot(dates, plays, marker='o')
        plt.xlabel('Date')
        plt.ylabel('Number of Plays')
        plt.title('Playback Trends Over Time')
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.show()

    def analyze_playback_reasons(self):
        start_reasons = Counter(info["reason_start"] for info in self.json_data)
        end_reasons = Counter(info["reason_end"] for info in self.json_data)

        results = "Playback start reasons:\n"
        for reason, count in start_reasons.items():
            results += f"{reason}: {count} times\n"

        results += "\nPlayback end reasons:\n"
        for reason, count in end_reasons.items():
            results += f"{reason}: {count} times\n"

        self.show_results(results)
        start_reasons = Counter(info["reason_start"] for info in self.json_data)
        end_reasons = Counter(info["reason_end"] for info in self.json_data)

        self.results_text.insert(tk.END, "Playback start reasons:\n")
        for reason, count in start_reasons.items():
            self.results_text.insert(tk.END, f"{reason}: {count} times\n")

        self.results_text.insert(tk.END, "\nPlayback end reasons:\n")
        for reason, count in end_reasons.items():
            self.results_text.insert(tk.END, f"{reason}: {count} times\n")

    def export_to_excel(self):
        if not self.json_data:
            self.results_text.insert(tk.END, "No JSON file loaded. Please open a JSON file first.\n")
            return

        if not self.current_json_file:
            self.results_text.insert(tk.END, "No current JSON file. Please open a JSON file first.\n")
            return

        excel_workbook = openpyxl.Workbook()
        sheet = excel_workbook.active
        sheet.title = "Spotify Statistics"

        current_row = 1
        sheet.cell(row=current_row, column=1, value="Date")
        sheet.cell(row=current_row, column=2, value="Hour")
        sheet.cell(row=current_row, column=3, value="Song")
        sheet.cell(row=current_row, column=4, value="Artist")
        sheet.cell(row=current_row, column=5, value="Minutes played")
        sheet.cell(row=current_row, column=6, value="Platform")
        sheet.cell(row=current_row, column=7, value="IP address")
        sheet.cell(row=current_row, column=8, value="Reason start")
        sheet.cell(row=current_row, column=9, value="Reason end")

        for info in self.json_data:
            current_row += 1
            ts = datetime.strptime(info["ts"], '%Y-%m-%dT%H:%M:%SZ')
            date = ts.strftime('%Y-%m-%d')
            hour = ts.strftime('%H:%M:%S')
            song = info["master_metadata_track_name"]
            artist = info["master_metadata_album_artist_name"]
            duration = info["ms_played"] / 1000 / 60
            formatted_duration = '{:.2f}'.format(duration)
            platform = info["platform"]
            ip_address = info["ip_addr_decrypted"]
            reason_start = info["reason_start"]
            reason_end = info["reason_end"]

            sheet.cell(row=current_row, column=1, value=date)
            sheet.cell(row=current_row, column=2, value=hour)
            sheet.cell(row=current_row, column=3, value=song)
            sheet.cell(row=current_row, column=4, value=artist)
            sheet.cell(row=current_row, column=5, value=formatted_duration)
            sheet.cell(row=current_row, column=6, value=platform)
            sheet.cell(row=current_row, column=7, value=ip_address)
            sheet.cell(row=current_row, column=8, value=reason_start)
            sheet.cell(row=current_row, column=9, value=reason_end)

        json_file_name = os.path.splitext(os.path.basename(self.current_json_file))[0]
        export_folder = os.path.dirname(self.current_json_file)
        excel_file = f"{json_file_name}_History.xlsx"
        excel_file_path = os.path.join(export_folder, excel_file)

        excel_workbook.save(excel_file_path)
    
        export_message = f"File '{json_file_name}' exported to '{excel_file_path}'.\n"
        
        self.results_text.insert(tk.END, export_message)
        self.show_results(export_message)
        
def main():
    root = tk.Tk()
    app = SpotifyAnalyzerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
